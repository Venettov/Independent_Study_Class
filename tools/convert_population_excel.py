#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Convert Census Excel -> clean CSV/JSON + compact PR series (2010–2024).

Enhancements vs. original:
- Reads TWO workbooks:
    A) prm-est2024-chg.xlsx (2020–2024)
    B) prm-est2020int-pop-72.xlsx (2010–2019)
- Same hardened header detection & robust numeric parsing.
- Cleans geographic names (removes "Municipio, Puerto Rico").
- Merges 2010–2019 + 2020–2024 into one wide table by municipio name.
- Overwrites prm-est2024-chg.csv/json with added 2010–2019 columns.
- Creates long-format series for ALL municipios (2010–2024).
- Extends Puerto Rico compact file to 2010–2024 (new filename).

Note: paths are relative to repo root (one level above this script).
"""

from __future__ import annotations
import json
import re
from pathlib import Path
from typing import Any, List, Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

# ---------- Paths ----------
REPO = Path(__file__).resolve().parents[1]

# New (2010–2019 intercensal)
XLSX_2010_2019 = REPO / "data" / "population" / "prm-est2020int-pop-72.xlsx"

# Existing (2020–2024)
XLSX_2020_2024 = REPO / "data" / "population" / "prm-est2024-chg.xlsx"

# Existing outputs (these will be OVERWRITTEN to include 2010–2019)
OUT_CSV_WIDE  = REPO / "data" / "population" / "prm-est2024-chg.csv"
OUT_JSON_WIDE = REPO / "data" / "population" / "prm-est2024-chg.json"

# New extras
OUT_SERIES_ALL_MUN = REPO / "data" / "population" / "pr_municipios_2010_2024.json"
OUT_PR_2010_2024   = REPO / "data" / "population" / "pr_pr_2010_2024.json"

def txt(v: Any) -> str:
    return "" if v is None else str(v).strip()

def norm(s: str) -> str:
    s = s.replace("\u2013", "-")  # en dash -> hyphen
    s = s.replace("\xa0", " ")    # NBSP -> space
    s = re.sub(r"\s+", " ", s).strip()
    return s

def uniquify(cols: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for c in cols:
        if c not in seen:
            seen[c] = 1
            out.append(c)
        else:
            out.append(f"{c}__{seen[c]}")
            seen[c] += 1
    return out

_num_keep = re.compile(r"[^0-9.\-]")

def to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, (int,)):
        return int(v)
    s = str(v)
    s = s.replace("\xa0", "")
    s = s.replace(",", "")
    s = _num_keep.sub("", s)
    if s in ("", "-", ".", "-.", ".-"):
        return None
    try:
        return int(float(s))
    except Exception:
        return None

_rm_municipio_suffix = re.compile(r"\s*Municipio,\s*Puerto\s*Rico\s*$", re.IGNORECASE)
def clean_geo_name(value: Any) -> Any:
    if value is None:
        return value
    s = str(value).strip()
    s = _rm_municipio_suffix.sub("", s)
    return s

def read_census_workbook(xlsx_path: Path, expected_years: List[int]) -> Tuple[pd.DataFrame, str]:
    """
    Read a Census-style workbook with parent/child headers, detect the geographic column,
    clean names, and return (df, geo_col).
    """
    print(f"\nReading workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    max_rows_scan = min(ws.max_row, 120)
    max_cols_scan = min(ws.max_column, 200)

    YEAR_SET = set(str(y) for y in expected_years)

    def is_year_token(cell_val: Any) -> bool:
        return txt(cell_val) in YEAR_SET

    # 1) Detect child header row
    child_hdr_row: Optional[int] = None
    for r in range(1, max_rows_scan + 1):
        hits = sum(1 for c in range(1, max_cols_scan + 1) if is_year_token(ws.cell(r, c).value))
        if hits >= max(3, min(5, len(expected_years)//2)):  # flexible threshold
            child_hdr_row = r
            break
    if child_hdr_row is None:
        raise RuntimeError(f"Couldn't detect child header row for {xlsx_path.name}")

    parent_hdr_row = max(1, child_hdr_row - 1)
    print(f"Header rows: parent={parent_hdr_row}, child={child_hdr_row}")

    # 2) Compose headers (parent forward-fill)
    parent_vals_raw = [norm(txt(ws.cell(parent_hdr_row, c).value)) for c in range(1, max_cols_scan + 1)]
    child_vals      = [norm(txt(ws.cell(child_hdr_row,  c).value)) for c in range(1, max_cols_scan + 1)]

    parent_vals: List[str] = []
    last_parent = ""
    for val in parent_vals_raw:
        if val:
            last_parent = val
            parent_vals.append(val)
        else:
            parent_vals.append(last_parent)

    def compose(parent: str, child: str, idx: int) -> str:
        if parent and child:
            return f"{parent} | {child}"
        if parent:
            return parent
        if child:
            return child
        return f"col_{idx}"

    headers = [compose(parent_vals[i-1], child_vals[i-1], i) for i in range(1, max_cols_scan + 1)]

    def column_is_empty(col_idx: int) -> bool:
        for r in range(child_hdr_row + 1, min(ws.max_row, child_hdr_row + 200) + 1):
            if txt(ws.cell(r, col_idx).value):
                return False
        return True

    while headers and headers[-1].startswith("col_") and column_is_empty(len(headers)):
        headers.pop()

    n_cols = len(headers)
    print(f"Detected {n_cols} header columns.")
    if n_cols:
        print("Columns sample:", headers[:12])

    # 3) Read data rows
    data_rows: List[List[Any]] = []
    for r in range(child_hdr_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, n_cols + 1)]
        if all(v in (None, "") for v in row_vals):
            break
        data_rows.append(row_vals)
    if not data_rows:
        raise RuntimeError(f"No data rows in {xlsx_path.name}")

    df = pd.DataFrame(data_rows, columns=uniquify(headers))

    # 4) Identify & clean geographic column
    geo_col: Optional[str] = None
    geo_regex = re.compile(r"(geographic.*area|area.*name|geographic.*name|\.geographic area)", re.I)
    for c in df.columns:
        if geo_regex.search(c):
            geo_col = c
            break
    if geo_col is None:
        # fallback: first column containing "Puerto Rico"
        for c in df.columns:
            s = df[c].astype(str).str.strip().str.lower()
            if s.str.contains(r"^puerto rico\b").any():
                geo_col = c
                break
    if geo_col is None:
        raise RuntimeError(f"Could not identify the 'Geographic Area' column in {xlsx_path.name}.")
    print(f"Geographic column: {geo_col}")

    df[geo_col] = df[geo_col].apply(clean_geo_name)

    return df, geo_col

def build_year_candidate_finder(df: pd.DataFrame) -> callable:
    """Returns candidate_columns_for_year(year) -> List[str] with flexible matching."""
    def find_exact(name: str) -> Optional[str]:
        for col in df.columns:
            if col.strip() == name:
                return col
        return None

    def find_contains(*parts: str) -> List[str]:
        parts_l = [p.lower() for p in parts]
        out = []
        for col in df.columns:
            cl = col.lower()
            if all(p in cl for p in parts_l):
                out.append(col)
        return out

    def candidate_columns_for_year(year: int) -> List[str]:
        cands: List[str] = []
        y = str(year)

        # 1) Exact child header that's just the year
        c = find_exact(y)
        if c: cands.append(c)

        # 2) Parent|child composed forms commonly seen
        #    Try a couple of likely parents used by Census spreadsheets
        for parent in [
            "Population Estimate (as of July 1)",
            "Intercensal Estimates",
            "Estimates", "Population",
        ]:
            composed = f"{parent} | {y}"
            c = find_exact(composed)
            if c and c not in cands: cands.append(c)

        # 3) Any column containing both "population" and the year
        for col in find_contains("population", y):
            if col not in cands: cands.append(col)

        # 4) Any column containing "estimate" and the year
        for col in find_contains("estimate", y):
            if col not in cands: cands.append(col)

        return cands
    return candidate_columns_for_year

def extract_years(df: pd.DataFrame, geo_col: str, years: List[int]) -> pd.DataFrame:
    """
    From a messy df, produce a clean wide table:
        [name, 2010, 2011, ..., 2024]
    """
    cand = build_year_candidate_finder(df)
    out = pd.DataFrame({ "name": df[geo_col].astype(str).str.strip() })

    for y in years:
        vals: List[Optional[int]] = []
        cands = cand(y)
        # Row-wise selection
        for _, row in df.iterrows():
            chosen_val: Optional[int] = None
            for col in cands:
                v = to_int(row.get(col, None))
                if v is not None:
                    chosen_val = v
                    break
            vals.append(chosen_val)
        out[str(y)] = vals
        print(f"Year {y}: candidate columns = {cands[:5]}{'...' if len(cands)>5 else ''}")

    return out

def merge_wide_tables(w2010_2019: pd.DataFrame, w2020_2024: pd.DataFrame) -> pd.DataFrame:
    """
    Left-join on 'name', coalescing year columns.
    If both have a year col (e.g., 2020 present in both), prefer non-null from 2020–2024 table.
    """
    left = w2010_2019.copy()
    right = w2020_2024.copy()

    merged = pd.merge(left, right, on="name", how="outer", suffixes=("_a", "_b"))

    # Coalesce all year columns 2010–2024
    for y in range(2010, 2025):
        ya, yb = f"{y}_a", f"{y}_b"
        if ya in merged.columns and yb in merged.columns:
            merged[str(y)] = merged[yb].combine_first(merged[ya])
            merged.drop([ya, yb], axis=1, inplace=True)
        elif ya in merged.columns:
            merged.rename(columns={ya: str(y)}, inplace=True)
        elif yb in merged.columns:
            merged.rename(columns={yb: str(y)}, inplace=True)
        else:
            # neither side had it (unlikely), create present but null
            merged[str(y)] = None

    # Keep only name + years in order
    ordered_cols = ["name"] + [str(y) for y in range(2010, 2025)]
    extra = [c for c in merged.columns if c not in ordered_cols]
    merged = merged[ordered_cols + extra]  # keep extras at the end for debugging
    return merged

def write_long_series_all_municipios(wide: pd.DataFrame, out_path: Path) -> None:
    """
    Emit [{"name": "...", "series":[{"year":2010,"population":...}, ...]}, ...]
    """
    records = []
    year_cols = [c for c in wide.columns if c.isdigit()]
    for _, row in wide.iterrows():
        series = []
        for y in sorted(int(c) for c in year_cols):
            val = row.get(str(y))
            val_int = None if pd.isna(val) else int(val)
            if val_int is not None:
                series.append({"year": y, "population": val_int})
        records.append({"name": row["name"], "series": series})

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote long series JSON for all municipios: {out_path}")

def main() -> None:
    # ---- Read both books
    df_2020_2024, geo_col_b = read_census_workbook(XLSX_2020_2024, list(range(2020, 2025)))
    df_2010_2019, geo_col_a = read_census_workbook(XLSX_2010_2019, list(range(2010, 2020)))

    # ---- Build wide tables from each
    wide_2010_2019 = extract_years(df_2010_2019, geo_col_a, list(range(2010, 2020)))
    wide_2020_2024 = extract_years(df_2020_2024, geo_col_b, list(range(2020, 2025)))

    # ---- Merge into one 2010–2024 table
    wide_merged = merge_wide_tables(wide_2010_2019, wide_2020_2024)

    # ---- Sort by name (Puerto Rico first, then others alpha)
    def sort_key(n: str) -> Tuple[int, str]:
        n = str(n)
        return (0 if n.strip().lower().startswith("puerto rico") else 1, n)
    wide_merged = wide_merged.sort_values(by="name", key=lambda s: s.map(sort_key)).reset_index(drop=True)

    # ---- Overwrite your existing wide CSV/JSON with added 2010–2019 columns
    OUT_CSV_WIDE.parent.mkdir(parents=True, exist_ok=True)
    wide_merged.to_csv(OUT_CSV_WIDE, index=False)
    wide_merged.to_json(OUT_JSON_WIDE, orient="records", force_ascii=False)
    print(f"Overwrote wide outputs with 2010–2024: {OUT_CSV_WIDE} | {OUT_JSON_WIDE}")

    # ---- Emit long series JSON for ALL municipios (2010–2024)
    write_long_series_all_municipios(wide_merged, OUT_SERIES_ALL_MUN)

    # ---- Puerto Rico compact series 2010–2024 (separate file)
    pr_mask = wide_merged["name"].astype(str).str.strip().str.lower().str.startswith("puerto rico")
    if not pr_mask.any():
        # try contains
        pr_mask = wide_merged["name"].astype(str).str.contains(r"\bpuerto rico\b", case=False, na=False)
    if not pr_mask.any():
        raise RuntimeError("Puerto Rico row not found after merge.")

    pr_row = wide_merged.loc[pr_mask].iloc[0]
    pr_series = []
    for y in range(2010, 2025):
        val = pr_row.get(str(y))
        if pd.notna(val):
            pr_series.append({"year": y, "population": int(val)})

    OUT_PR_2010_2024.write_text(
        json.dumps({"name": "Puerto Rico", "series": pr_series}, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"Wrote Puerto Rico compact series 2010–2024: {OUT_PR_2010_2024}")

if __name__ == "__main__":
    main()
