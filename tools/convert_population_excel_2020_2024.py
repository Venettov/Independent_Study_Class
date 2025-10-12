#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Convert Census Excel -> clean CSV/JSON + compact PR series (2020–2024).

Hardened version:
- Autodetects child header row by exact year tokens.
- Parent = child - 1; forward-fills across merged spans.
- Composes unique headers "Parent | Child".
- Robust numeric parsing (strip commas, NBSP, any non-digits except '-' and '.').
- For each year, tries multiple candidate columns in priority order.
- NEW: Cleans geographic names by removing the suffix "Municipio, Puerto Rico".
"""

from __future__ import annotations
import json
import re
from pathlib import Path
from typing import Any, List, Dict, Optional

import pandas as pd
from openpyxl import load_workbook

# ---------- Paths ----------
REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "data" / "population" / "prm-est2024-chg.xlsx"
OUT_CSV = REPO / "data" / "population" / "prm-est2024-chg.csv"
OUT_JSON = REPO / "data" / "population" / "prm-est2024-chg.json"
OUT_PR   = REPO / "data" / "population" / "pr_pr_2020_2024.json"

def txt(v: Any) -> str:
    return "" if v is None else str(v).strip()

def norm(s: str) -> str:
    s = s.replace("\u2013", "-")  # en dash -> hyphen
    s = s.replace("\xa0", " ")    # NBSP -> space
    s = re.sub(r"\s+", " ", s).strip()
    return s

def uniquify(cols: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for c in cols:
        if c not in seen:
            seen[c] = 1
            out.append(c)
        else:
            out.append(f"{c}__{seen[c]}")
            seen[c] += 1
    return out

_num_keep = re.compile(r"[^0-9.\-]")  # remove everything except digits, dot, minus

def to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, (int,)):
        return int(v)
    s = str(v)
    s = s.replace("\xa0", "")  # NBSP
    s = s.replace(",", "")
    s = _num_keep.sub("", s)  # strip footnotes/letters/units
    if s in ("", "-", ".", "-.", ".-"):
        return None
    try:
        return int(float(s))
    except Exception:
        return None

# NEW: cleaner for geographic names
_rm_municipio_suffix = re.compile(r"\s*Municipio,\s*Puerto\s*Rico\s*$", re.IGNORECASE)
def clean_geo_name(value: Any) -> Any:
    if value is None:
        return value
    s = str(value).strip()
    # Only remove the exact suffix "Municipio, Puerto Rico"
    s = _rm_municipio_suffix.sub("", s)
    return s

print(f"Reading workbook: {XLSX}")
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]

max_rows_scan = min(ws.max_row, 120)
max_cols_scan = min(ws.max_column, 200)

# --- 1) Detect child header row: row with at least 3 cells that are exactly one of the year tokens
YEAR_SET = {"2020", "2021", "2022", "2023", "2024"}

def is_year_token(cell_val: Any) -> bool:
    return txt(cell_val) in YEAR_SET

child_hdr_row: Optional[int] = None
for r in range(1, max_rows_scan + 1):
    hits = sum(1 for c in range(1, max_cols_scan + 1) if is_year_token(ws.cell(r, c).value))
    if hits >= 3:
        child_hdr_row = r
        break

if child_hdr_row is None:
    raise RuntimeError("Couldn't detect the child header row (no row with multiple exact year labels).")

parent_hdr_row = max(1, child_hdr_row - 1)
print(f"Header rows: parent={parent_hdr_row}, child={child_hdr_row}")

# --- 2) Build parent + child headers and forward-fill parent across merged spans
parent_vals_raw = [norm(txt(ws.cell(parent_hdr_row, c).value)) for c in range(1, max_cols_scan + 1)]
child_vals      = [norm(txt(ws.cell(child_hdr_row,  c).value)) for c in range(1, max_cols_scan + 1)]

parent_vals: List[str] = []
last_parent = ""
for val in parent_vals_raw:
    if val:
        last_parent = val
        parent_vals.append(val)
    else:
        parent_vals.append(last_parent)

def compose(parent: str, child: str, idx: int) -> str:
    if parent and child:
        return f"{parent} | {child}"
    if parent:
        return parent
    if child:
        return child
    return f"col_{idx}"

headers = [compose(parent_vals[i-1], child_vals[i-1], i) for i in range(1, max_cols_scan + 1)]

# Trim trailing empty columns (no data below)
def column_is_empty(col_idx: int) -> bool:
    for r in range(child_hdr_row + 1, min(ws.max_row, child_hdr_row + 200) + 1):
        if txt(ws.cell(r, col_idx).value):
            return False
    return True

while headers and headers[-1].startswith("col_") and column_is_empty(len(headers)):
    headers.pop()

n_cols = len(headers)
print(f"Detected {n_cols} header columns.")
print("Columns sample:", headers[:12])

# --- 3) Read data rows until first fully-empty row
data_rows: List[List[Any]] = []
for r in range(child_hdr_row + 1, ws.max_row + 1):
    row_vals = [ws.cell(r, c).value for c in range(1, n_cols + 1)]
    if all(v in (None, "") for v in row_vals):
        break
    data_rows.append(row_vals)
if not data_rows:
    raise RuntimeError("No data rows found under headers.")

# Build DataFrame with unique columns
df = pd.DataFrame(data_rows, columns=uniquify(headers))

# --- 4) Identify and CLEAN the geographic column BEFORE saving outputs
geo_col = None
geo_regex = re.compile(r"(geographic.*area|area.*name|geographic.*name|\.geographic area)", re.I)
for c in df.columns:
    if geo_regex.search(c):
        geo_col = c
        break
if geo_col is None:
    # fallback: first column containing "Puerto Rico"
    for c in df.columns:
        s = df[c].astype(str).str.strip().str.lower()
        if s.str.contains(r"^puerto rico\b").any():
            geo_col = c
            break
if geo_col is None:
    raise RuntimeError("Could not identify the 'Geographic Area' column.")
print(f"Geographic column: {geo_col}")

# NEW: apply the cleaner
df[geo_col] = df[geo_col].apply(clean_geo_name)

# --- 5) Save full table (now with cleaned names)
OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
df.to_csv(OUT_CSV, index=False)
df.to_json(OUT_JSON, orient="records", force_ascii=False)
print("Wrote table CSV and JSON.")

# --- 6) Puerto Rico row (unchanged by cleaning)
pr_mask = df[geo_col].astype(str).str.strip().str.lower().str.startswith("puerto rico")
if not pr_mask.any():
    pr_mask = df[geo_col].astype(str).str.contains(r"\bpuerto rico\b", case=False, na=False)
if not pr_mask.any():
    raise RuntimeError("Could not find a row for 'Puerto Rico'.")
pr_row = df.loc[pr_mask].iloc[0]

# --- 7) Helpers to find candidate columns for each year
def find_exact(name: str) -> Optional[str]:
    for col in df.columns:
        if col.strip() == name:
            return col
    return None

def find_contains(*parts: str) -> List[str]:
    parts_l = [p.lower() for p in parts]
    out = []
    for col in df.columns:
        cl = col.lower()
        if all(p in cl for p in parts_l):
            out.append(col)
    return out

def candidate_columns_for_year(year: int) -> List[str]:
    """Priority list of possible columns for a given year."""
    cands = []
    # 1) Exact composed header
    exact = f"Population Estimate (as of July 1) | {year}"
    c = find_exact(exact)
    if c: cands.append(c)
    # 2) Contains "Population Estimate" + year
    cands.extend([col for col in find_contains("population estimate", str(year)) if col not in cands])
    # 3) Plain year header (e.g., "2020")
    c = find_exact(str(year))
    if c and c not in cands: cands.append(c)
    # 4) Special fallback for 2020 base
    if year == 2020:
        base = find_exact("April 1, 2020 Estimates Base")
        if not base:
            base_list = find_contains("april 1, 2020", "base")
            base = base_list[0] if base_list else None
        if base and base not in cands:
            cands.append(base)
    return cands

def get_year_value(row: pd.Series, year: int) -> tuple[Optional[int], Optional[str], Any]:
    """Try all candidate columns; return (int_value, chosen_col, raw_value)"""
    for col in candidate_columns_for_year(year):
        raw = row.get(col, None)
        val = to_int(raw)
        if val is not None:
            return val, col, raw
    return None, None, None

# --- 8) Build compact JSON series with robust fallback
series = []
chosen_map: Dict[int, str] = {}
for y in [2020, 2021, 2022, 2023, 2024]:
    val, chosen_col, raw = get_year_value(pr_row, y)
    if val is None:
        print(f"[DEBUG] Could not parse a value for {y}. Candidates were: {candidate_columns_for_year(y)}")
        raise RuntimeError(f"Null/invalid value for {y}")
    series.append({"year": y, "population": val})
    chosen_map[y] = chosen_col
    print(f"Year {y}: using column '{chosen_col}' with raw='{raw}' -> {val}")

OUT_PR.write_text(json.dumps({"name": "Puerto Rico", "series": series}, ensure_ascii=False, indent=2), encoding="utf-8")
print("Wrote compact series JSON:", OUT_PR)
print("Chosen columns:", chosen_map)
print("All done.")
